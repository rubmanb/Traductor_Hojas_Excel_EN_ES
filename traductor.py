
import openpyxl
from translate import Translator

# Cargar el archivo Excel
archivo_excel = "grocery-shopping-list.xlsx" #-> nombre del fichero xlsx (Excel)
hoja_excel = "Grocery List" #-> nombre de la hoja del fichero xlsx que queremos traducir

# Abrir el archivo Excel
libro_excel = openpyxl.load_workbook(archivo_excel)
hoja = libro_excel[hoja_excel]

# Recorrer las celdas con los nombres en inglés y traducirlos al español
for fila in hoja.iter_rows(min_row=1, min_col=1, max_row=hoja.max_row, max_col=hoja.max_column):
    for celda in fila:
        # Obtener el texto en inglés de la celda
        texto_ingles = celda.value
        
        try:
            # Traducir el texto al español
            traductor = Translator(to_lang="es", from_lang="en")
            texto_espanol = traductor.translate(texto_ingles)
        except:
            # Omitir errores de traducción y conservar el texto original
            texto_espanol = texto_ingles
        
        # Asignar el texto traducido a la misma celda
        celda.value = texto_espanol

# Guardar los cambios en el archivo Excel
libro_excel.save(archivo_excel)

# Cerrar el archivo Excel
libro_excel.close()

print("Traducción completada.")
